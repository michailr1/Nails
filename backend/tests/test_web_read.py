from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime, timedelta
from decimal import Decimal

from conftest import WEB_ORIGIN_HEADERS
from openpyxl import load_workbook
from sqlalchemy import select

from app.config import get_settings
from app.db import get_session_factory
from app.models import AuditEvent, Booking, BookingStatus, Client, ClientProfileStatus, Service
from app.services.normalization import normalize_public_name
from app.services.web_auth import _keyed_hash
from app.web_auth_models import WebSession


def _authenticate(client, user_id: uuid.UUID) -> None:
    now = datetime.now(UTC)
    token = "session-token-" + str(user_id)
    settings = get_settings()
    with get_session_factory()() as session:
        session.add(
            WebSession(
                token_hash=_keyed_hash(
                    token,
                    purpose="session-token",
                    settings=settings,
                ),
                user_id=user_id,
                last_seen_at=now,
                idle_expires_at=now + timedelta(hours=1),
                absolute_expires_at=now + timedelta(days=1),
                rotation_counter=1,
                created_ip_hash="a" * 64,
                last_ip_hash="a" * 64,
                request_id="web-read-test",
            )
        )
        session.commit()
    client.cookies.set("__Host-nails_session", token)


def _seed_owner_data(owner_user_id: uuid.UUID, *, suffix: str) -> tuple[Client, Booking]:
    starts_at = datetime(2026, 7, 20, 9, 0, tzinfo=UTC)
    with get_session_factory()() as session:
        client = Client(
            owner_user_id=owner_user_id,
            public_name=f"=Клиентка {suffix}",
            normalized_public_name=normalize_public_name(f"=Клиентка {suffix}"),
            phone=f"+49000{suffix}",
            notes=f"private-{suffix}",
            profile_status=ClientProfileStatus.active,
        )
        service = Service(
            owner_user_id=owner_user_id,
            public_name=f"Маникюр {suffix}",
            normalized_public_name=normalize_public_name(f"Маникюр {suffix}"),
            price_amount=Decimal("2500.00"),
            currency="RUB",
            duration_minutes=120,
            buffer_before_minutes=0,
            buffer_after_minutes=20,
            is_active=True,
        )
        session.add_all([client, service])
        session.flush()
        booking = Booking(
            owner_user_id=owner_user_id,
            client_id=client.id,
            service_id=service.id,
            starts_at=starts_at,
            ends_at=starts_at + timedelta(hours=2),
            reserved_starts_at=starts_at,
            reserved_ends_at=starts_at + timedelta(hours=2, minutes=20),
            duration_minutes_snapshot=120,
            buffer_before_minutes_snapshot=0,
            buffer_after_minutes_snapshot=20,
            status=BookingStatus.scheduled,
            price_amount=Decimal("2500.00"),
            currency="RUB",
            price_source="service_snapshot",
            idempotency_key=f"web-read-{suffix}",
        )
        session.add(booking)
        session.commit()
        session.refresh(client)
        session.refresh(booking)
        session.expunge(client)
        session.expunge(booking)
        return client, booking


def test_web_read_requires_session(client, clean_database):
    response = client.get(
        "/web/api/clients",
        headers=WEB_ORIGIN_HEADERS,
    )
    assert response.status_code == 401


def test_calendar_and_clients_are_owner_scoped(client, create_user):
    owner = create_user(telegram_user_id=100000001)
    other = create_user(telegram_user_id=100000002)
    owner_client, owner_booking = _seed_owner_data(owner.id, suffix="owner")
    _seed_owner_data(other.id, suffix="other")
    _authenticate(client, owner.id)

    clients = client.get("/web/api/clients", headers=WEB_ORIGIN_HEADERS)
    assert clients.status_code == 200
    payload = clients.json()["clients"]
    assert len(payload) == 1
    assert payload[0]["client_id"] == str(owner_client.id)
    assert payload[0]["notes"] == "private-owner"

    calendar = client.get(
        "/web/api/calendar",
        headers=WEB_ORIGIN_HEADERS,
        params={"date_from": "2026-07-20", "date_to": "2026-07-20"},
    )
    assert calendar.status_code == 200
    bookings = calendar.json()["bookings"]
    assert len(bookings) == 1
    assert bookings[0]["booking_id"] == str(owner_booking.id)


def test_calendar_range_is_limited(client, create_user):
    owner = create_user()
    _authenticate(client, owner.id)
    response = client.get(
        "/web/api/calendar",
        headers=WEB_ORIGIN_HEADERS,
        params={"date_from": "2026-01-01", "date_to": "2026-03-01"},
    )
    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "date_range_too_large"


def test_csv_export_sanitizes_formulas_and_audits_without_content(
    client,
    create_user,
):
    owner = create_user()
    _seed_owner_data(owner.id, suffix="owner")
    _authenticate(client, owner.id)

    response = client.post(
        "/web/api/exports/clients",
        headers=WEB_ORIGIN_HEADERS,
        params={"format": "csv"},
    )
    assert response.status_code == 200
    text = response.content.decode("utf-8-sig")
    assert "'=Клиентка owner" in text
    assert "private-owner" in text

    with get_session_factory()() as session:
        event = session.scalar(
            select(AuditEvent).where(AuditEvent.action == "web_export_created")
        )
        assert event is not None
        assert event.owner_user_id == owner.id
        assert event.safe_changes == {"format": "csv", "row_count": 1}
        assert "private-owner" not in str(event.safe_changes)


def test_xlsx_export_sanitizes_formula_cells(client, create_user):
    owner = create_user()
    _seed_owner_data(owner.id, suffix="owner")
    _authenticate(client, owner.id)

    response = client.post(
        "/web/api/exports/clients",
        headers=WEB_ORIGIN_HEADERS,
        params={"format": "xlsx"},
    )
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    worksheet = workbook["Клиентки"]
    assert worksheet.cell(row=2, column=1).value == "'=Клиентка owner"
